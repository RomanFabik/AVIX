import streamlit as st
import pandas as pd
from deep_translator import GoogleTranslator
import time
import io
from openpyxl import load_workbook
from openpyxl.styles import Font
import base64
import re

# === AVIX SETTINGS ===
PRIMARY_GREEN = "#275437"
DARK_BACKGROUND = "#232323"
LIGHT_TEXT = "#EEEEEE"
WHITE = "#FFFFFF"
FONT_URL = "https://fonts.googleapis.com/css2?family=Roboto+Mono&display=swap"

# === LANGUAGE CODE MAP ===
# AVIX / older Google codes -> deep-translator supported codes
LANG_MAP = {
    "in": "id",      # Indonesian
    "iw": "he",      # Hebrew
    "fil": "tl",     # Filipino / Tagalog
    "zh": "zh-CN",   # Chinese Simplified
}

def normalize_lang_code(code):
    code = str(code).strip()
    return LANG_MAP.get(code, code)

def translate_batch_safe(texts, source_lang, target_lang, chunk_size=50):
    """Translate a list of texts in chunks. If target language fails, fallback to English."""
    source_lang = normalize_lang_code(source_lang)
    target_lang = normalize_lang_code(target_lang)

    result = ["" for _ in texts]
    items = [(i, str(text).strip()) for i, text in enumerate(texts) if str(text).strip()]

    if not items:
        return result, set()

    fallback_indices = set()

    for start in range(0, len(items), chunk_size):
        chunk = items[start:start + chunk_size]
        idxs = [i for i, _ in chunk]
        chunk_texts = [text for _, text in chunk]

        try:
            translated = GoogleTranslator(source=source_lang, target=target_lang).translate_batch(chunk_texts)
            if translated is None:
                translated = ["" for _ in chunk_texts]
        except Exception:
            try:
                translated = GoogleTranslator(source=source_lang, target="en").translate_batch(chunk_texts)
                translated = [f"[FALLBACK EN] {t}" if t else "" for t in translated]
                fallback_indices.update(idxs)
            except Exception as e:
                translated = [f"[CHYBA] {str(e)}" for _ in chunk_texts]
                fallback_indices.update(idxs)

        for i, value in zip(idxs, translated):
            result[i] = "" if value is None else str(value)

    return result, fallback_indices

st.set_page_config(page_title="AVIX AI Translation", page_icon=":earth_africa:", layout="wide")

# === TRANSLATIONS ===
translations = {
    "sk": {
        "upload_file": "Nahraj XLSX alebo XLS súbor",
        "select_column": "Vyber zdrojový stĺpec (napr. Slovak (sk))",
        "source_language": "Zdrojový jazyk (napr. sk, en)",
        "select_target": "Vyber cieľové jazyky",
        "translate_button": "Preložiť",
        "preview_translation": "Náhľad prekladaného hárku",
        "preview_result": "Náhľad preloženého hárku",
        "download_file": "📥 Stiahnuť preložený XLSX súbor",
        "success_translation": "Preklad dokončený za {seconds:.2f} sekúnd.",
    },
    "en": {
        "upload_file": "Upload XLSX or XLS file",
        "select_column": "Select source column (e.g., Slovak (sk))",
        "source_language": "Source language (e.g., sk, en)",
        "select_target": "Select target languages",
        "translate_button": "Translate",
        "preview_translation": "Preview of translation sheet",
        "preview_result": "Preview of translated sheet",
        "download_file": "📥 Download translated XLSX file",
        "success_translation": "Translation completed in {seconds:.2f} seconds.",
    },
    "de": {
        "upload_file": "XLSX oder XLS-Datei hochladen",
        "select_column": "Quellspalte auswählen (z.B. Slovak (sk))",
        "source_language": "Ausgangssprache (z.B. sk, en)",
        "select_target": "Zielsprachen auswählen",
        "translate_button": "Übersetzen",
        "preview_translation": "Vorschau des Übersetzungsblatts",
        "preview_result": "Vorschau des übersetzten Blatts",
        "download_file": "📥 Übersetzte XLSX-Datei herunterladen",
        "success_translation": "Übersetzung abgeschlossen in {seconds:.2f} Sekunden.",
    }
}

# === STYLES ===
st.markdown(f"""
    <style>
        @import url('{FONT_URL}');
        html, body, [class*="css"] {{
            font-family: 'Roboto Mono', monospace;
            background-color: {DARK_BACKGROUND};
            color: {LIGHT_TEXT};
        }}
        .stButton>button, .stDownloadButton>button {{
            background-color: {PRIMARY_GREEN};
            color: white;
            font-weight: bold;
        }}
        footer {{ visibility: hidden; }}
    </style>
""", unsafe_allow_html=True)

# === LOGO ===
def load_logo_base64(path):
    with open(path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()

logo_base64 = load_logo_base64("avix_logo.png")

# === HEADER & LANGUAGE ===
col_header, col_lang = st.columns([5, 1])
with col_header:
    st.markdown(f"""
        <div style="display:flex;align-items:center;justify-content:space-between;">
            <div style="display:flex;align-items:center;gap:1rem;">
                <img src="data:image/png;base64,{logo_base64}" height="50">
                <h1>AVIX AI Translation</h1>
            </div>
            <a href="https://www.avix.eu" style="color:{PRIMARY_GREEN};font-weight:bold;">www.avix.eu</a>
        </div>
    """, unsafe_allow_html=True)

with col_lang:
    lang_choice = st.selectbox("🌐", ["sk", "en", "de"], format_func=lambda x: {"sk": "🇸🇰", "en": "🇬🇧", "de": "🇩🇪"}[x])

t = translations[lang_choice]

# === UPLOAD ===
col1, col2 = st.columns([1, 4])
with col1:
    st.write(t["upload_file"])
with col2:
    uploaded_file = st.file_uploader(t["upload_file"], type=["xlsx", "xls"], label_visibility="collapsed")

# === PROCESSING ===
if uploaded_file:
    try:
        xls_bytes = uploaded_file.read()
        file_name = uploaded_file.name.lower()
        if file_name.endswith(".xls"):
            # starý Excel formát – potrebuješ mať nainštalované `xlrd`
            xls = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=None, engine="xlrd")
        else:
            # .xlsx – ako doteraz
            xls = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=None, engine="openpyxl")
        translation_df = xls[list(xls.keys())[0]]
        configuration_df = xls[list(xls.keys())[1]]

        with st.expander(t["preview_translation"], expanded=True):
            st.dataframe(translation_df.head())

        lang_col_pattern = re.compile(r".*\(([\w-]{2,10})\)")
        candidate_cols = {
            col: translation_df[col].notna().sum()
            for col in translation_df.columns
            if lang_col_pattern.match(col)
        }

        auto_text_column = max(candidate_cols, key=candidate_cols.get) if candidate_cols else translation_df.columns[0]
        auto_source_lang = lang_col_pattern.match(auto_text_column).group(1) if candidate_cols else "sk"

        c1, c2, c3 = st.columns([2, 2, 3])
        with c1:
            text_column = st.selectbox(t["select_column"], translation_df.columns, index=translation_df.columns.get_loc(auto_text_column))
        with c2:
            source_lang = st.text_input(t["source_language"], auto_source_lang)
        with c3:
            lang_col_pattern = re.compile(r".*\(([\w-]{2,10})\)")
            existing_target_langs = []
            
            for col in translation_df.columns:
                match = lang_col_pattern.match(col)
                if match:
                    lang_code = match.group(1)
                    if lang_code != source_lang:
                        existing_target_langs.append(lang_code)

            # všetky jazykové kódy, ktoré sú v XLS (okrem zdrojového)
            all_lang_options = sorted(set(existing_target_langs))

            target_langs = st.multiselect(
                t["select_target"],
                all_lang_options,
                default=all_lang_options  # predvolene označí všetky dostupné jazyky
            )


        col_btn = st.columns([1, 6, 1])[1]
        with col_btn:
            if st.button(t["translate_button"]):
                start_time = time.time()

                # Dôležité pre Streamlit Cloud / novšie pandas:
                # prázdne Excel stĺpce sa načítajú ako float64, preto ich pred zápisom
                # prepíšeme na čisté Python object/string hodnoty.
                translation_df_copy = translation_df.copy().astype("object")
                translation_df_copy = translation_df_copy.where(pd.notna(translation_df_copy), "")

                total_rows = len(translation_df_copy)
                progress_bar = st.progress(0)
                cell_styles = {}
                suspicious_words = ['poloz', 'rama', 'skrutky', 'ulozenie']

                for lang in target_langs:
                    matching_col = next((col for col in translation_df_copy.columns if str(col).lower().endswith(f"({lang})")), None)
                    if not matching_col:
                        matching_col = f"Translation ({lang})"
                        translation_df_copy[matching_col] = pd.Series([""] * len(translation_df_copy), dtype="object")
                    else:
                        translation_df_copy[matching_col] = translation_df_copy[matching_col].astype("object")

                source_texts = translation_df[text_column].fillna("").astype(str).str.strip().tolist()
                total_jobs = max(1, len(target_langs))

                for lang_i, lang in enumerate(target_langs, start=1):
                    matching_col = next((col for col in translation_df_copy.columns if str(col).lower().endswith(f"({lang})")), None)
                    if not matching_col:
                        matching_col = f"Translation ({lang})"
                        translation_df_copy[matching_col] = pd.Series([""] * len(translation_df_copy), dtype="object")

                    translated_values, fallback_indices = translate_batch_safe(
                        source_texts,
                        source_lang=source_lang,
                        target_lang=lang,
                        chunk_size=50,
                    )

                    translation_df_copy.loc[:, matching_col] = pd.Series(
                        translated_values,
                        index=translation_df_copy.index,
                        dtype="object"
                    )

                    for pos, translated_text in enumerate(translated_values):
                        if pos in fallback_indices or str(translated_text).startswith("[CHYBA]"):
                            cell_styles[(translation_df_copy.index[pos], matching_col)] = "highlight"
                        elif translated_text and any(word.lower() in translated_text.lower() for word in suspicious_words):
                            cell_styles[(translation_df_copy.index[pos], matching_col)] = "highlight"

                    progress_bar.progress(lang_i / total_jobs)

                st.success(t["success_translation"].format(seconds=time.time() - start_time))

                with st.expander(t["preview_result"], expanded=True):
                    st.dataframe(translation_df_copy.head())

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    translation_df_copy.to_excel(writer, sheet_name='Translations', index=False)
                    configuration_df.to_excel(writer, sheet_name='Configuration', index=False)
                output.seek(0)

                wb = load_workbook(output)
                ws = wb['Translations']

                from openpyxl.styles import Font

                # Nastav Arial 10 pre všetky bunky v preklade
                default_font = Font(name="Arial", size=10)

                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None:
                            cell.font = default_font

                ws_config = wb['Configuration']
                default_font = Font(name="Arial", size=10)

                for row in ws_config.iter_rows(min_row=1, max_row=ws_config.max_row, min_col=1, max_col=ws_config.max_column):
                    for cell in row:
                        if cell.value is not None:
                            cell.font = default_font


                for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                    letter = col_cells[0].column_letter
                    ws.column_dimensions[letter].width = 80 if letter != "A" else 1

                index_to_excel_row = {idx_value: pos + 2 for pos, idx_value in enumerate(translation_df_copy.index)}
                for (row_idx, col_name), _ in cell_styles.items():
                    col_idx = list(translation_df_copy.columns).index(col_name) + 1
                    excel_row = index_to_excel_row.get(row_idx, row_idx + 2 if isinstance(row_idx, int) else 2)
                    ws.cell(row=excel_row, column=col_idx).font = Font(color="FF0000", bold=True)

                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)

                st.download_button(
                    label=t["download_file"],
                    data=final_output,
                    file_name="preklad.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as e:
        st.error(f"Chyba pri spracovaní súboru: {e}")
